#Instruction 1
#importation pour les routes
from flask import Flask, render_template, send_from_directory, url_for, flash, redirect, session, request, logging
import shutil
#importation pour mysql
# from flask_sqlalchemy import SQLAlchemy
# from flask_login import UserMixin
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import numpy as np
import pandas.io.sql
import pyodbc
from flask_mysqldb import MySQL
from werkzeug.utils import secure_filename
from functools import wraps
from wtforms import Form, SelectField, TextAreaField, PasswordField, SubmitField, MultipleFileField, StringField, DateField, validators
from passlib.hash import sha256_crypt
from pathlib import Path
import mysql.connector
import datetime
import os
from datetime import datetime, date
import time
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
#import schedule
import threading
from apscheduler.schedulers.background import BackgroundScheduler

#Instruction 2
app = Flask(__name__)
app.config['SECRET_KEY'] = 'supersecretkey'

#MySQL config
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''

app.config['MYSQL_DB'] = 'autorisation' #'basicBD'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

#init MySQL
mysql = MySQL(app)

def is_logged_in(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            flash('Vous n\'êtes pas autorisé à acceder à ce compte, veuillez vous connecté d\'abord!!!', 'danger')
            return redirect(url_for('login'))  
    return wrap 

@app.route('/')
def acceuil():
    return render_template('acceuil.html') 

@app.route('/page')
def page():
    return render_template('page.html') 

class UploadFileForm(Form):

    instance = SelectField ( "Serveur", choices=[("10.10.120.40\pamecas_SQL"), ("10.10.120.41"), ("10.10.120.42")] )  
    caisse = SelectField ( "Caisse", choices=[("DEFAULT"),("vdn"), ("AMBAO"), ("amecbt"), ("bayakh"), ("MECZON"), ("MECTG"), 
                                              ("DAHRA"), ("BOURGUIBA"), ("FATICK"), ("diaobe"), ("cdgy"), ("MECIB"), ("MECTHIES"), 
                                              ("TASSETTE"), ("TOUBATOUL"), ("TOUBA"), ("MECNI"), ("DIOURBEL"), ("MECREL"), ("YOFF")] ) 

    title = StringField ('Titre', [validators.input_required()])
    comment = TextAreaField("Commentaire")
    
    file = MultipleFileField("Fichiers", [validators.input_required()])
    submit = SubmitField("Ajouter")


def connect_to_Sqlserver():
        
        colonne = ['COD_EMPRESA','AGENCE','NUM_CLIENT','NOM_CLIENT','NUMERO_CREDIT','DATE_D\'OUVERTURE','MONTANT_DEBOURSE',
                'SOLDE_CREDIT','ETAT','COD_ESTADO_COMPTABLE','DES_ESTADO_COMPTABLE','FEC_VENCIMIENTO','COD_ID','NUM_ID']
        df = pd.DataFrame(columns = colonne)
        for serveur in ("SVR1", "SVR2", "SVR3"):
            conn = pyodbc.connect(
                "Driver={SQL Server};Server="+serveur+";UID=test;PWD=test;Database=reporting;")
  
            df1 = pd.read_sql_query("""    
                SELECT
                    [COD_EMPRESA]
                    ,[AGENCE]
                    ,[NUM_CLIENT]
                    ,[NOM_CLIENT] 
                    ,[NUMERO_CREDIT]
                    ,[DATE_D'OUVERTURE]
                    ,[MONTANT_DEBOURSE]
                    ,[SOLDE_CREDIT]
                    ,[ETAT]
                    ,[COD_ESTADO_COMPTABLE]
                    ,[DES_ESTADO_COMPTABLE]
                    ,[FEC_VENCIMIENTO]
                    ,[COD_ID]
                    ,[NUM_ID]
                FROM [REPORTING].[dbo].[CEIC]                                                      
                                        """, conn)           
            
            df = pd.concat([df, df1])

        df.to_csv ('ceic1.csv')

        df2 = pd.read_csv('ceic1.csv')
        df2 = df2["AGENCE"].drop_duplicates()
        df2 = pd.DataFrame(df)
        df2 = df2.dropna()

        df3 = pd.read_csv('caisses.csv')
        df3 = df3[~df3['AGENCE'].isin(df2['AGENCE'])]
        df3.to_csv('bnc.csv')

        df4 = pd.read_csv('bnc.csv')
        df4 = df4.loc[:, ~df4.columns.str.contains('^Unnamed')]
        df4.set_index('AGENCE', inplace=True)
        #df4 = df4.dropna()
        df4.to_excel ('Bases_non_chargees.xlsx')
        os.remove('bnc.csv')

        if not df4.empty :

            sender = "pbmbaye@pamecas.sn"
            receiver = "pbmbaye@pamecas.sn" #"mbfall@pamecas.sn" #[owner,"supportdsi@pamecas.sn"] 
            port = 465  # For SSL
            password = "SanFranCisco2022"
            txt = (
                f"Bonjour,\n"
                f"Merci de recevoir la liste des bases non chargées.\n"
                f"Cordialement."                              
                  )
            msg = MIMEMultipart()
            msg ['sender'] = "Papa Badara MBAYE"
            msg['Subject'] = "CEIC" 
            msg.attach(MIMEText(txt))
            fileattach = 'Bases_non_chargees.xlsx'
            with open(fileattach, "rb") as fil:
                    part = MIMEApplication(
                        fil.read()
                )
            part['Content-Disposition'] = 'attachment; filename="%s"' % fileattach
            msg.attach(part)

            context = ssl.create_default_https_context = ssl._create_unverified_context()
            ssl._create_default_https_context = ssl._create_unverified_context
            with smtplib.SMTP_SSL("smtp.pamecas.sn", port, context=context) as server:
                server.login("pbmbaye@pamecas.sn", password)
                server.sendmail(sender, receiver, msg.as_string())  

            os.remove('Bases_non_chargees.xlsx') 

        else:
            sender = "pbmbaye@pamecas.sn"
            receiver = "pbmbaye@pamecas.sn"  #[owner,"supportdsi@pamecas.sn"]  #"mbfall@pamecas.sn"
            port = 465  # For SSL
            password = "SanFranCisco2022"            
            txt = (
                f"Bonjour,\n"
                f"Votre demande est traitée.\n"
                f"Cordialement...!!!"
                  )
            msg = MIMEMultipart()
            msg ['sender'] = "Papa Badara MBAYE"
            msg['Subject'] = "CEIC" 
            msg.attach(MIMEText(txt))
            fileattach = 'Bases_non_chargees.xlsx'
            with open(fileattach, "rb") as fil:
                    part = MIMEApplication(
                        fil.read()
                )
            part['Content-Disposition'] = 'attachment; filename="%s"' % fileattach
            msg.attach(part)

            context = ssl.create_default_https_context = ssl._create_unverified_context()
            ssl._create_default_https_context = ssl._create_unverified_context
            with smtplib.SMTP_SSL("smtp.pamecas.sn", port, context=context) as server:
                server.login("pbmbaye@pamecas.sn", password)
                server.sendmail(sender, receiver, msg.as_string()) 

            if os.path.isfile('ceic.csv'):
                os.remove('ceic.csv')
            os.rename('ceic1.csv','ceic.csv')   
        conn.commit()
        conn.close() 

scheduler = BackgroundScheduler()
scheduler.add_job(func=connect_to_Sqlserver, trigger="cron", minute=10, second=00)
scheduler.start()
# #scheduler.remove_all_jobs()

        
@app.route('/ceic', methods = ['GET', 'POST'])
@is_logged_in

def ceic():
    form = UploadFileForm(request.form)
    if request.method == 'POST': # and form.validate():

        #serveur = request.form['instance']
        database  = request.form['caisse']  
        file = request.files['file']

        file.save( os.path.join ( file.filename ) )

        # df = pd.read_excel ( 'ceic.xlsx' ) 
        df = pd.read_csv ( 'ceic.csv' ) 
        df.set_index('COD_EMPRESA', inplace=True)
        
        df3 = pd.read_excel ( file.filename ) 
        df3 = df3.applymap(str) 

        df3 = df[ df["NUM_ID"].isin( df3["NUM_ID"] ) ]
        df3 = df3.loc[:, ~df3.columns.str.contains('^Unnamed')]

        #now = date.today( )
        now = time.localtime()
        Date = time.strftime('%d-%m-%Y', now)
        now = datetime.now()

        df3.to_excel('ceic '+database+' '+str(Date)+'.xlsx')

        a = 'ceic '+database+' '+str(Date)+'.xlsx'

        # wb = load_workbook('ceic '+database+' '+str(Date)+'.xlsx')
        # ws = wb['Sheet1']

        # for letter in ['C','D','E','F','G','H','I','J','K','L','M','N']:
            
        #     max_width = 0

        #     for row_number in range (1, ws.max_row + 1):

        #         if len(ws[f'{letter}{row_number}'].value) > max_width:

        #             max_width = len(ws[f'{letter}{row_number}'].value)

        #     ws.column_dimensions[letter].width = max_width + 1

        # wb.save('ceic '+database+' '+str(Date)+'.xlsx')

        cur = mysql.connection.cursor()

        cur.execute("INSERT INTO ceic (email, caisse,  Date) VALUES(%s,%s,%s)" , 
                   ( session['email'] , database,  now))

        cur.execute ("SELECT * FROM ceic WHERE email = %s", [session['email']])
        demande = cur.fetchone()

        #name = demande['name']
        # owner = demande['email']

        # sender = "pbmbaye@pamecas.sn"
        # receiver = owner #[owner,"supportdsi@pamecas.sn"] 
        # port = 465  # For SSL
        # password = "SanFranCisco2022"
        # txt = (
        #     f"Bonjour,\n"
        #     f"Votre demande est traitée.\n"
        #     f"Cordialement."
        #       )
        # msg = MIMEMultipart()
        # msg ['sender'] = "Papa Badara MBAYE"
        # msg['Subject'] = "CEIC" 
        # msg.attach(MIMEText(txt))
        # fileattach = 'ceic '+database+' '+str(Date)+'.xlsx'
        # with open(fileattach, "rb") as fil:
        #         part = MIMEApplication(
        #             fil.read()
        #     )
        # part['Content-Disposition'] = 'attachment; filename="%s"' % fileattach
        # msg.attach(part)

        # context = ssl.create_default_https_context = ssl._create_unverified_context()
        # ssl._create_default_https_context = ssl._create_unverified_context
        # with smtplib.SMTP_SSL("smtp.pamecas.sn", port, context=context) as server:
        #     server.login("pbmbaye@pamecas.sn", password)
        #     server.sendmail(sender, receiver, msg.as_string())   
           
        directory = r'C:\Users\Badara\Documents\CEIC\Archives'
        #directory = r'C:\Users\Badara\Videos\Archives'

        if not os.path.exists(directory):

            os.makedirs(directory)  
        
        if os.path.exists( os.path.join ( directory, file.filename ) ): #Love

            os.rename(file.filename, 'new_'+file.filename)
            shutil.move('new_'+file.filename, directory)

        else:

            shutil.move(file.filename, directory)

        directory1 = r'C:\Users\Badara\Documents\CEIC\Demandes'

        if not os.path.exists(directory1):

            os.makedirs(directory1)  

        shutil.move(a, directory1)

        flash('La demande a bien été envoyée', 'success')

        #return redirect(url_for('ceic'))
        return redirect(url_for('download_demandes'))    
        #return redirect(url_for('table', filename = a)) 
        return redirect(url_for('download_demandes'))                 
       
    return render_template('ceic.html', form = form)
 
    #return render_template('table.html', name=a, data=df3.to_html())     
    #return render_template('ceic.html' ,form=form, tables=df3.to_html(), titles=df3.columns.values)   


@app.route('/charger', methods = ['GET', 'POST'])
def charger():
    form = UploadFileForm(request.form)
    if request.method == 'POST':    
        colonne = ['COD_EMPRESA','AGENCE','NUM_CLIENT','NOM_CLIENT','NUMERO_CREDIT','DATE_D\'OUVERTURE','MONTANT_DEBOURSE',
            'SOLDE_CREDIT','ETAT','COD_ESTADO_COMPTABLE','DES_ESTADO_COMPTABLE','FEC_VENCIMIENTO','COD_ID','NUM_ID']
        df = pd.DataFrame(columns = colonne)
        for serveur in ("10.10.120.40\pamecas_SQL", "10.10.120.41", "10.10.120.42"):
            conn = pyodbc.connect(
                "Driver={SQL Server};Server="+serveur+";UID=test;PWD=test;Database=reporting;")
  
            df1 = pd.read_sql_query("""    
                SELECT
                    [COD_EMPRESA]
                    ,[AGENCE]
                    ,[NUM_CLIENT]
                    ,[NOM_CLIENT]
                    ,[NUMERO_CREDIT]
                    ,[DATE_D'OUVERTURE]
                    ,[MONTANT_DEBOURSE]
                    ,[SOLDE_CREDIT]
                    ,[ETAT]
                    ,[COD_ESTADO_COMPTABLE]
                    ,[DES_ESTADO_COMPTABLE]
                    ,[FEC_VENCIMIENTO]
                    ,[COD_ID]
                    ,[NUM_ID]
                FROM [REPORTING].[dbo].[CEIC]                                                      
                                        """, conn)           
            
            df = pd.concat([df, df1])

        df.to_csv ('ceic1.csv')

        if os.path.isfile('ceic.csv'):
            os.remove('ceic.csv')
        os.rename('ceic1.csv','ceic.csv')   
        conn.commit()
        conn.close() 
        #flash('Le chargement s\'est effectué avec succés ', 'success')
    #return redirect(url_for('charger'))
    return render_template('charger.html', form = form)

@app.route('/charger1', methods = ['GET', 'POST'])
def charger1():
    form = UploadFileForm(request.form)
    if request.method == 'POST':    
 
        df = pd.read_csv('ceic.csv')
        df.set_index('COD_EMPRESA', inplace=True)

        df1 = df.loc[df['AGENCE'].isin(['MECTHIES', 'AGENCE MECTHIES','NOTTO/TASSETTE','MECTOUBATOUL','MEC/NDIAYENE SIRAKH','MEC/NGOUNDIANE'])] 
        # df2 = df.loc[df['AGENCE'].isin(['PAMECAS-TOUBA', 'POINT DE SERVICE DE MBACKE','PAMECAS DIOURBEL','PAMECAS BAMBEY'])]
        # df3 = df.loc[df['AGENCE'].isin(['PAMECAS DAHRA'])]

        df1 = df1.loc[:, ~df1.columns.str.contains('^Unnamed')]
        df1.to_excel ('AgenceThies.xlsx')
        # df2 = df1.loc[:, ~df1.columns.str.contains('^Unnamed')]
        # df2.to_excel ('AgenceThies.xlsx')
        # df3 = df1.loc[:, ~df1.columns.str.contains('^Unnamed')]
        # df3.to_excel ('AgenceThies.xlsx')

        df1 = df1 [['AGENCE','DES_ESTADO_COMPTABLE','SOLDE_CREDIT']]
        pivot_table = df1.pivot_table(index='AGENCE', columns='DES_ESTADO_COMPTABLE',values='SOLDE_CREDIT', aggfunc='sum')
        pivot_table.to_excel('pivot_AgenceThies.xlsx', 'Rapport')

        wb = load_workbook('pivot_AgenceThies.xlsx')
        sheet = wb['Rapport']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        barchart = BarChart()

        data = Reference ( sheet, min_col = min_column+1, max_col = max_column, min_row = min_row, max_row = max_row )
        categories = Reference ( sheet, min_col = min_column, max_col = min_column, min_row = min_row+1, max_row = max_row )

        barchart.add_data(data, titles_from_data = True)
        barchart.set_categories(categories)

        sheet.add_chart(barchart, 'A10' )

        barchart.title = "Credits par caisses"
        barchart.style = 10
        wb.save ('Graphique_barre.xlsx') 

        #Send Mail

        os.remove('AgenceThies.xlsx')
        os.remove('pivot_AgenceThies.xlsx')

        flash('Le chargement s\'est effectué avec succés ', 'success')
    return render_template('charger1.html', form = form)


class RegisterForm(Form):
    nom = SelectField('Nom', [validators.input_required()])
    email = SelectField('Email', [validators.input_required()])
    password  = PasswordField('Mot de passe', 
        [        
        validators.DataRequired(),
        validators.EqualTo('confirm', message='Les mots de passes ne sont pas les mêmes')
        ])
    confirm = PasswordField('Confirmer le mot de passe')   


@app.route('/register', methods=['GET', 'POST'])         
def register():

    form = RegisterForm(request.form)
    if request.method == 'POST' and form.validate():
        name = form.nom.data
        email = form.email.data
        password = sha256_crypt.encrypt(str(form.password.data))

        #creer curseur
        now = datetime.now()
        ts = time.mktime(datetime.strptime(str(now), '%Y-%m-%d %H:%M:%S.%f').timetuple())
        Date = datetime.fromtimestamp(ts)
        cur = mysql.connection.cursor()

        cur.execute("INSERT INTO users (name, email, password, date_register) VALUES(%s,%s,%s,%s)" , (name, email, password, Date))

        mysql.connection.commit()
        cur.close()
        flash('Bonjour ' +name+ ', votre compte est créé avec succes ', 'success')
        return redirect(url_for('login'))
        #return render_template('register.html')
    return render_template('register.html', form=form)
    

@app.route('/login', methods = ['GET', 'POST'] )
def login():

    if request.method == 'POST':
        email = request.form['email']
        password_candidate = request.form['password']

        cur =   mysql.connection.cursor()
        cur1 =  mysql.connection.cursor()
        cur2 =  mysql.connection.cursor()
        cur3 =  mysql.connection.cursor()

        #get user by email
        result =   cur.execute  ("SELECT *  FROM user      WHERE email = %s", [email])
        result1 =  cur1.execute ("SELECT *  FROM admin     WHERE email = %s", [email])
        result2 =  cur2.execute ("SELECT *  FROM users     WHERE email = %s", [email])
        result3 =  cur3.execute ("SELECT *  FROM staff_dcp WHERE email = %s", [email])

        
            
        if result > 0 and result1 > 0 and result == result1: 
            data = cur.fetchone()
            password = data['password']
            name = data['name']
            #compare password
            if sha256_crypt.verify(password_candidate, password):
                #app.logger.info('PASSWORD MATCHED')
                session['logged_in'] = True
                session['email'] = email

                flash('Bienvenue ' +name+ '', 'success')
                return redirect(url_for('charger'))
                #return render_template('ajouter_demande.html', form = form)    
            else:
                flash( 'Les paramètres de connexions ne sont pas valides, Merci reéssayer', 'danger')
                return redirect(url_for('login'))
            
        elif result > 0 and result3 > 0 and result2 > 0 and  result == result2 == result3: 
            data = cur.fetchone()
            password = data['password']
            name = data['name']
            #compare password
            if sha256_crypt.verify(password_candidate, password):
                #app.logger.info('PASSWORD MATCHED')
                session['logged_in'] = True
                session['email'] = email

                flash('Bienvenue ' +name+ '', 'success')
                return redirect(url_for('charger1'))
                #return render_template('ajouter_demande.html', form = form)    
            else:
                flash( 'Les paramètres de connexions ne sont pas valides, Merci reéssayer', 'danger')
                return redirect(url_for('login'))
            
        elif result > 0 and result2 > 0 and  result == result2: 
            data = cur.fetchone()
            password = data['password']
            name = data['name']
            #compare password
            # if not sha256_crypt.verify(password_candidate, password):
            #     print("Locked")
            # else:
            #     print("Access Granted")
            if sha256_crypt.verify(password_candidate, password):
                #app.logger.info('PASSWORD MATCHED')
                session['logged_in'] = True
                session['email'] = email

                flash('Bienvenue ' +name+ '', 'success')
                return redirect(url_for('ceic'))
                #return render_template('ajouter_demande.html', form = form)    
            else:
                flash( 'Les paramètres de connexions ne sont pas valides, Merci reéssayer', 'danger')
                return redirect(url_for('login'))   
            
        else: 
            flash( 'Cet email n\'a pas de compte, Merci d\'en créer ', 'danger')            
            return redirect(url_for('register')) 

        cur.close() 
    return render_template('login.html')
        

@app.route('/logout')
def logout():
    session.clear()
    flash('Vous êtes maintenant deconnecté', 'success')
    return redirect(url_for('login'))

@app.route('/table/<filename>')
def table(filename):
    directory = r'C:\Users\Badara\Documents\CEIC\Demandes'

    if not os.path.exists(directory):
        os.makedirs(directory)      

    if os.path.exists( os.path.join ( directory, filename ) ): #Love
        df = pd.read_excel (os.path.join ( directory, filename ))

    return render_template ("table.html", name='Resulte de la demande CEIC', data=df.to_html())

@app.route('/download_Demandes')
def download_demandes():
    return render_template('download_demandes.html', files = os.listdir(r'C:\Users\Badara\Documents\CEIC\Demandes') )

@app.route('/uploads/<filename>')
def download_demandesfile(filename):
    return send_from_directory(r'C:\Users\Badara\Documents\CEIC\Demandes', filename)

#Instruction 4
if __name__ == '__main__':
    
    app.secret_key='$€cret_87'
    #app.run()  #app.run sans argument, on est obligé de redemarrer le serveur à chaque modif du fichier   
    #threading.Thread(target=connect_to_Sqlserver).start() #Permet d'executer la fonction connect_to_Sqlserver tout le temps en arriere plan
    app.run(debug='True')

#'cron', day_of_week='mon-fri', hour=5, minute=30
# schedule.every().day.at("14:42").do(connect_to_Sqlserver)
# while True:
#     schedule.run_pending()
#     time.sleep(1)




